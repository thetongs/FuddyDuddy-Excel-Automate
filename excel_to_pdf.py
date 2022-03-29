# Libraries
from win32com import client
import win32api
import pathlib
import os
import openpyxl

# Excel to PDF Automate
def excel_conversion(excel_path_source, excel_path_destination):
    base_file_name = os.path.basename(excel_path_source).split('.')[0] 
    excel_path = str(pathlib.Path.cwd() / excel_path_source)
    book = openpyxl.load_workbook(excel_path_source)

    # Create Result Folder
    try:
        os.mkdir(excel_path_destination + "/{}_PDF Files".format(base_file_name))
    except Exception:
        print("Folder issue")
 
    # For Each Sheet
    for sheet_number, sheet_name in zip(range(3, len(book.sheetnames) + 1), book.sheetnames[2:]):
        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0
        wb = excel.Workbooks.Open(excel_path)

        ws = wb.Worksheets[sheet_number,sheet_number].Select()
        excel_path_destination1 = "/{}_PDF Files".format(base_file_name) + '/' + "{}.pdf".format(sheet_name)
        new_add = excel_path_destination + excel_path_destination1
        pdf_path = str(pathlib.Path.cwd() / new_add)

        try:
            wb.SaveAs(pdf_path, FileFormat = 57)
        except Exception as e:
            print(e)
        finally:
            wb.Close()
            excel.Quit()